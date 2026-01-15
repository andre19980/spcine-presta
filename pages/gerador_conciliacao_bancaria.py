import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from functions.get_range import get_range_conciliacao_bancaria
from functions.utils import generate_excel_file

from layout.user_menu import user_menu

def gerador_conciliacao_bancaria():
  st.title("Gerador de Conciliação Bancária")
  user_menu()

  uploaded_file = st.file_uploader("Escolha um arquivo .xlsx", type=["xlsx"])

  if uploaded_file is not None:
    handle_file(uploaded_file)

  return

def handle_file(file):
  wb = load_workbook(file, data_only=True)

  st.divider()

  with st.container(horizontal=True, horizontal_alignment='center'):
    st.image("logo_spcine-principal.png", width=200)

  df = pd.DataFrame({})

  for sheet in wb.sheetnames:
    ws = wb[sheet]

    [START, END] = get_range_conciliacao_bancaria("dt. balancete", "999 s a l d o", ws)
    HEADER_ROW, LAST_ROW = START[0], END[0]
    START_COL, LAST_COL = START[1] - 1, END[1]
    nrows = LAST_ROW - HEADER_ROW

    df2 = pd.read_excel(
      file,
      sheet_name=sheet,
      skiprows=HEADER_ROW - 1,
      header=0,
      nrows=nrows,
      usecols=list(range(START_COL, LAST_COL)),
    )

    df_cleaned = df2.loc[:, ~df2.columns.str.contains('^Unnamed')]
    df = pd.concat([df, df_cleaned], ignore_index=True)

  df['group_id'] = df['Dt. balancete'].notnull().cumsum()

  df_result = df.groupby('group_id', as_index=False).agg({
    'Dt. balancete': 'first',
    'Dt. movimento': 'first',
    'Ag. origem': 'first',
    'Lote': 'first',
    'Histórico': lambda x: ' '.join(x.dropna().astype(str)),
    'Documento': 'first',
    'Valor R$': 'first',
    'Saldo': 'first'
  })

  df_result['Ag. origem'] = df_result['Ag. origem'].fillna(0).astype(int).astype(str).str.zfill(4)
  df_result['Lote'] = df_result['Lote'].fillna(0).astype(int).astype(str).str.zfill(5)

  df_result = df_result.drop(columns='group_id')

  df_values = df_result[df_result['Valor R$'].notna()].copy()
  df_values['Marcador'] = df_values['Valor R$'].str[-1]
  df_values['Valor Absoluto'] = pd.to_numeric(df_values['Valor R$'].str[:-1].str.replace('.', '').str.replace(',', '.'))
  df_values['Valor Numérico'] = np.where(df_values['Marcador'] == 'D', -df_values['Valor Absoluto'], df_values['Valor Absoluto'])

  df_values = df_values.drop(columns=['Marcador', 'Valor Absoluto'])
  df_selected_columns = df_values[['Dt. balancete', 'Histórico', 'Valor R$', 'Valor Numérico']].copy()

  df_selected_columns['Aplicações / Resgates'] = np.where(
    df_selected_columns['Histórico'].str.contains('791', na=False),
    df_selected_columns['Valor Numérico'],
    None
  )

  filtro_pagamentos = ~df_selected_columns['Histórico'].str.contains(r'435|791', na=False)
  df_selected_columns['Pagamentos'] = np.where(
    filtro_pagamentos,
    df_selected_columns['Valor Numérico'],
    None
  )

  df_selected_columns['Tarifas Bancárias'] = np.where(
    df_selected_columns['Histórico'].str.contains('435', na=False),
    df_selected_columns['Valor Numérico'],
    None
  )

  st.table(df_selected_columns)

  generate_excel_file(df_selected_columns)
